import pandas as pd
import re
import traceback
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

# === Helper Functions ===
def clean_header_name(raw_text):
    text = re.sub(r'Ch\d+\s*', '', str(raw_text))
    text = re.sub(r'MP\s*\d+', '', text)
    return text.strip()

def fill_merged_range(sheet, start_row, start_col, end_row, end_col, text, fill_color="FFFF00"):
    align_center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    top_left_cell = sheet.cell(row=start_row, column=start_col)
    top_left_cell.value = text
    top_left_cell.fill = header_fill
    top_left_cell.alignment = align_center

def unmerge_and_rename_cooling(ws_dat, merged_cell):
    cooling_headers = [
        "High Ambient/Idle/DB Notch Solenoid Control",
        "Low Ambient/Low HP Solenoid Control",
        "Low Ambient/Medium HP Solenoid Control",
        "Low Ambient/High HP Solenoid Control",
        "High Ambient/Idle/DB Notch Fan Control",
        "Low Ambient/Low HP Fan Control",
        "Low Ambient/Medium HP Fan Control",
        "Low Ambient/High HP Fan Control",
        "HIGH AMBIENT/HIGH NOTCH SOLENOID CONTROL",
        "HIGH AMBIENT/LOW NOTCH SOLENOID CONTROL",
        "LOW AMBIENT/HIGH NOTCH SOLENOID CONTROL",
        "LOW AMBIENT/LOW NOTCH SOLENOID CONTROL",
        "HIGH AMBIENT/HIGH NOTCH FAN CONTROL",
        "HIGH AMBIENT/LOW NOTCH FAN CONTROL",
        "LOW AMBIENT/HIGH NOTCH FAN CONTROL",
        "LOW AMBIENT/LOW NOTCH FAN CONTROL"
    ]

    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    ws_dat.unmerge_cells(str(merged_cell))

    for i, header_name in enumerate(cooling_headers):
        col_index = merged_cell.min_col + i
        ws_dat.cell(row=1, column=col_index).value = header_name
        ws_dat.cell(row=1, column=col_index).fill = blue_fill

def enable_wrap_text_first_row(ws_dat):
    for col in range(1, ws_dat.max_column + 1):
        cell = ws_dat.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def safe_read_par(par_file):
    processed_lines = []
    try:
        content = par_file.read().decode("utf-8", errors="ignore").splitlines()
        for line in content:
            line = line.strip()
            if not line:
                continue
            parts = line.split("|")
            trimmed_parts = [p.strip()[:20] for p in parts]
            processed_lines.append(trimmed_parts)

        # Limit to 11 rows max (rows 6–16)
        processed_lines = processed_lines[:11]

        max_cols = max(len(row) for row in processed_lines)
        normalized_data = [row + [""] * (max_cols - len(row)) for row in processed_lines]
        col_names = [f"Col_{i+1}" for i in range(max_cols)]
        df = pd.DataFrame(normalized_data, columns=col_names)
        return df
    except Exception as e:
        st.error(f"Error reading PAR file: {e}")
        return pd.DataFrame()

def process_par_headers(ws_dat, ws_par):
    sizes = {"F": 16, "B": 8, "W": 1, "U": 1, "P": 1, "N": 1}
    next_start_col = 2

    for row in ws_par.iter_rows(min_row=1, max_row=ws_par.max_row):
        first_col_value = str(row[0].value).strip() if row[0].value else ""
        if not first_col_value:
            continue

        first_letter = first_col_value[0].upper()
        line_text = str(row[-1].value).strip() if row[-1].value else ""
        if "|" in line_text:
            label = line_text.split("|")[-1].strip()
        else:
            label = line_text
        label = clean_header_name(label)

        if first_letter in sizes:
            size = sizes[first_letter]
            if label.lower() == "cooling status":
                fill_merged_range(ws_dat, 1, next_start_col, 1, next_start_col + size - 1, label)
                for merged_cell in ws_dat.merged_cells.ranges:
                    if merged_cell.min_row == 1 and merged_cell.min_col == next_start_col:
                        unmerge_and_rename_cooling(ws_dat, merged_cell)
                        break
            else:
                fill_merged_range(ws_dat, 1, next_start_col, 1, next_start_col + size - 1, label)
            next_start_col += size

# === Streamlit Web App ===
st.title("📊 DAT & PAR to Excel Converter")

dat_file = st.file_uploader("Upload .dat file", type=["dat"])
par_file = st.file_uploader("Upload .par file", type=["par"])

if st.button("Run Conversion"):
    if not dat_file or not par_file:
        st.error("Please upload both .dat and .par files.")
    else:
        try:
            df_par = safe_read_par(par_file)
            excel_file = "output.xlsx"

            with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
                # Write placeholder DAT sheet
                placeholder_headers = ["Time"] + [f"Col{i}" for i in range(2, 101)]
                pd.DataFrame(columns=placeholder_headers).to_excel(writer, index=False, sheet_name="DAT_Data")

                # Write PAR data starting from row 6 (index=5)
                df_par.to_excel(writer, index=False, sheet_name="PAR_Data", startrow=5)

            wb = load_workbook(excel_file)
            ws_dat = wb["DAT_Data"]
            ws_par = wb["PAR_Data"]

            process_par_headers(ws_dat, ws_par)
            wb.save(excel_file)

            df_dat = pd.read_csv(dat_file, sep=None, engine="python", header=None)
            if not df_dat.empty:
                df_dat.columns = ["Time"] + [f"Col{i}" for i in range(2, len(df_dat.columns) + 1)]

            wb = load_workbook(excel_file)
            ws_dat = wb["DAT_Data"]

            for r_idx, row in enumerate(df_dat.values.tolist(), start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws_dat.cell(row=r_idx, column=c_idx).value = value

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws_dat["A1"].value = "Time"
            ws_dat["A1"].fill = yellow_fill

            enable_wrap_text_first_row(ws_dat)
            wb.save(excel_file)

            with open(excel_file, "rb") as f:
                st.success("✅ Excel file generated successfully!")
                st.download_button("📥 Download Excel File", f, file_name="output.xlsx")

        except Exception:
            st.error(f"An error occurred:\n{traceback.format_exc()}")
